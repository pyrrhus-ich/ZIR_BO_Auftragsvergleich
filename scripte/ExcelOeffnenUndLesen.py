import openpyxl as op
from openpyxl.styles import PatternFill, Font
from colorama import *
init(autoreset=True)



# liest das src File und speichert die Werte als Liste in einer Variablen ab
# src = Sourcefile | dst = Variable die die Werte speichert
def readSrc(src, dst, maxColumn):
    print("Auslesen src File beginnt : readSrc()")
    wb = op.load_workbook(src,data_only=True) # lädt das File
    ws = wb.worksheets[0]
    #schreibt die Werte in eine Liste wenn der erste value der Zeile nicht none ist
    for value in ws.iter_rows(min_row=1, min_col=1,max_col=maxColumn, values_only=True):
            if value[0] is not None:
                dst.append(list(value)) # Ganz wichtig das die values als Liste gespeichert werden, da sonst unveränderlich
    print(Fore.GREEN + "Auslesen Source File ist beendet : readSrc()")
    



 

