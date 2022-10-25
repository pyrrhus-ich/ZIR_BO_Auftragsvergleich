from scripte.WWS_Retoure.Variablen_WWSRetoure import *
import openpyxl as op
from openpyxl.styles import PatternFill, Font
from colorama import *
init(autoreset=True)




def createDstFile(dstFileName, baseList): #dstFile, resulList
    print("Beginne jetzt mit dem schreiben des dst Files createDstFile()")
    wbDst = op.Workbook() # erzeugt Workbook Objekt mit einem Sheet
    ws = wbDst.active     # aktiviert das erste sheet
    numberOfColumns = len(baseList[0]) # Legt die Anzahl der Spalten fest die es gibt | MUSS Noch in Buchstaben übersetzt werden
    for row in baseList: #Hängt jeden Eintrag in der resultList an das neu erzeugt sheet
        ws.append(row)
# Hier versuche ich die Spaltenbreite zu ermitteln um diese dann zu setzen
    """
    columnA = 0
    for el in baseList:
        if len(el[0]) > columnA:
            columnA =len(el[0])
    """
    #<<<<<<<<<< Hier beginnt die Formatierung des Excel Files >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    print("Wir starten das schön machen des Files")
    
    # Setzt die Spaltenbreiten
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 35
    
    ws.auto_filter.ref = ws.dimensions  #Setzt den Autofilter über alle Spalten
   
    # Setze für die Überschrift den Background Color zu Blau | Schriftgröße zu 14 | Schriftfarbe zu Weiss
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col= numberOfColumns):
        for cell in rows:
            cell.fill = PatternFill(start_color="2f48ff", end_color="2f48ff", fill_type = "solid")
            cell.font = Font(size=14, color="FFFFFF")
       
    wbDst.save(dstFileName) #Speichert das File
    wbDst.close()
    print(Fore.GREEN + "Das schreiben des Ausgabefiles ist abgeschlossen. Es wird bereitgestellt in : {}\n".format(dstFileName))   
    print(Style.RESET_ALL, end="")#Setzt die Farbeinstellungen wieder zurück
