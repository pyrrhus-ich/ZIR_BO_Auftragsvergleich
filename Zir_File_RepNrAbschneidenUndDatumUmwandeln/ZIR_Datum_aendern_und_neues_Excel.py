import openpyxl as op
from openpyxl import load_workbook
import os
import pdb


workDir = os.getcwd()           
srcFld=workDir + "\\srcZir\\"    
srcFile = srcFld + os.listdir(srcFld)[0]  
srcFileName=os.listdir(srcFld)[0]  
dstFile = workDir +"\\dstZir\\changed-"+ srcFileName

print("srcFileName " + srcFileName)
print("srcFile " + srcFile)
print("dstFile " + dstFile)

"""
Wichtig die folgenden Variablen bezeichnet den Index der Spalte in der das "Date of creation" und die Zir Reparatur
auftragsnummer ("M199-12345-R") steht
Index fängt bei 0 an zu zählen. Also ist 2 die 3. Spalte
"""
indRepairNumber= 0
indDateOfCreation = 2

print("Index des Reparaturauftrages = " + indRepairNumber)
print("Index des 'Date of creation' = " + indDateOfCreation)

#pdb.set_trace() # setzt hier den Breakpoint

print("Starte_Verarbeitung")
wb=load_workbook(filename=srcFile)
ws=wb.worksheets[0]

print("Objekt definiert")
def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]
print("iter_row_ definiert")
val = []

print("Beginne Variable 'val' zu füllen")

for el in iter_rows(ws):
    val.append(el)

print("Insert in val abgeschlossen")
print("Es wurden {} Datensätze (ohne Kopfzeile) eingelesen".format(len(val)-1))

val[0].insert(indDateOfCreation + 1,"Tag")
val[0].insert(indDateOfCreation + 2,"Monat")
val[0].insert(indDateOfCreation + 3,"Jahr")

print("Neue Spalten in Val eingefügt")

y=""
jahr = ""
monat =""
tag = ""

print("Beginne mit Umwandeln des Datums")

for el in val:
    if el != val[0]:
        #Hier bearbeiten wir die Auftragsnummer und schneiden die letzten 2 "-R" oder mehr ab
        i = (el[indRepairNumber])
        slice_object = slice(0,10)
        el0 = i[slice_object]
        el.pop(indRepairNumber)
        el.insert(indRepairNumber,el0)
        #Bearbeitung der Auftragsnummer beendet
        # Hier wird das Datum bearbeitet
        z = el[2].split(" ")
        x=z[0].split("-")
        x.reverse()
        tag = x[0]
        monat = x [1]
        jahr=x[2]
        x.insert(1,".")
        x.insert(3,".")
        for i in x:
            y=y+i
        el.pop(indDateOfCreation)
        el.insert(indDateOfCreation,y)
        el.insert(indDateOfCreation + 1, tag)
        el.insert(indDateOfCreation + 2, monat)
        el.insert(indDateOfCreation + 3, jahr)
    y=""

print("Abschneiden der Reparaturnummer & Datumsumwandlung abgeschlossen")

def createDstFile(baseList, dstFileName): #resulList, dstFile
    print("Beginne jetzt mit dem schreiben des dst Files createDstFile()")
    wbDst = op.Workbook() # erzeugt Workbook Objekt mit einem Sheet
    ws = wbDst.active     # aktiviert das erste sheet
    for row in baseList: #Hängt jeden Eintrag in der resultList an das neu erzeugt sheet
        ws.append(row)
    ws.auto_filter.ref = ws.dimensions          #Setzt den Autofilter über alle Spalten
    wbDst.save(dstFileName) #Speichert das File
    wbDst.close()
    print("Das schreiben des Ausgabefiles ist abgeschlossen")   
   
"""
Hier wird die Funktion zum erstellen des dstFiles aufgerufen
"""
createDstFile(val, dstFile)
