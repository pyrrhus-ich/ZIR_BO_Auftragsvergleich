import openpyxl as op
from openpyxl.styles import PatternFill, Font
from colorama import init, Fore, Style # Colorama für farbige Terminalausgaben
from basic import *
init()



# liest das src File und speichert die Werte als Liste in einer Variablen ab
def readSrc(sourceFile, dataList):
    print("Auslesen src File beginnt : readSrc()")
    wb = op.load_workbook(sourceFile,data_only=True) # lädt das File
    ws = wb.worksheets[0]
    #schreibt die Werte in eine Liste wenn der erste value der Zeile nicht none ist
    for value in ws.iter_rows(min_row=2, min_col=1,max_col=1, values_only=True):
            #if value[1] is not None:
                #dataList.append(value)
            dataList.append(list(value))
    print("Auslesen Source File ist beendet : readSrc()")
    
# Liest die in readSrc() erstellte Liste aus und speichert die Werte in unterschiedlichen Variablen
def createSingleLists(srcList, list0, list1):
    for el in valList:
        if el[0] is not None:
            list0.append(el[0])
        if el[1] is not None:
            list1.append(el[1])



 

